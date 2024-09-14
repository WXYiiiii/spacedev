import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sqlite3

# 读取配置文件
def load_config(config_path='config.yaml'):
    with open(config_path, 'r', encoding='utf-8') as file:
        return yaml.safe_load(file)

# 连接SQLite，执行查询并获取结果
def query_sqlite_data(conn, project_id, sqlite_indicators, config):
    cursor = conn.cursor()
    query = f"""
    SELECT {', '.join(sqlite_indicators)}
    FROM project_data
    WHERE {config['mapping']['project_id_column']} = ?
    """
    cursor.execute(query, (project_id,))        # 根据维度的id进行查询  -> 每一行的数据可以对应上
    return cursor.fetchone()

# 设置单元格样式的函数    1. 字体   2. 背景   3. 位置
def set_cell_style(cell, config):
    cell.font = Font(
        name=config['excel_style']['font']['name'],
        size=config['excel_style']['font']['size'],
        bold=config['excel_style']['font']['bold'],
        italic=config['excel_style']['font']['italic'],
        color=config['excel_style']['font']['color']
    )
    cell.fill = PatternFill(start_color=config['excel_style']['fill']['start_color'])
    cell.alignment = Alignment(
        horizontal=config['excel_style']['alignment']['horizontal'],
        vertical=config['excel_style']['alignment']['vertical']
    )

# 读取: Excel模板       获取: 1. 指标名称 2. 维度名称和行数的映射关系
def read_excel_template(input_path, config):
    """
    读取指定的Excel模板，提取项目名称及其对应的行号，并获取指标名称的映射关系。

    参数:
    - template_path (str): Excel模板的文件路径。
    - config (dict): 包含映射和样式配置的字典。

    返回:
    - workbook: 加载的Excel工作簿对象。
    - sheet: Excel工作表对象。
    - excel_indicators (list): 从Excel中提取的指标名称列表。
    - project_map (dict): 项目名称到行号的映射字典。
    - indicator_column_map (dict): 指标名称到其列索引的映射字典。
    """

    workbook = openpyxl.load_workbook(input_path)
    sheet = workbook.active

    # 指标字段:  获取Excel中的指标字段，从指定列(维度列的下一列)开始
    excel_indicators = [cell.value for cell in sheet[config['mapping']['indicator_row']][config['mapping']['column_num']:]]

    # 维度和行号的映射: 获取维度名称和对应的Excel行号   //excel中的行号和列号都是从1开始
    # 字典推导式: {key: value for item in iterable if condition}
    project_map = {
        sheet[f"{config['mapping']['project_column']}{row}"].value: row
        for row in range(config['mapping']['start_row'], sheet.max_row + 1)
        if sheet[f"{config['mapping']['project_column']}{row}"].value
    }
    # print(project_map)

    return workbook, sheet, excel_indicators, project_map

# 将查询结果写入Excel
def write_to_excel(sheet, project_row, query_result, excel_to_sqlite_indicator_map, config):
    """
    将查询结果写入指定的Excel工作表。

    参数:
    - sheet: Excel工作表对象，表示要写入数据的工作表。
    - project_row (int): 当前项目在Excel中的行号。
    - query_result (tuple): 从SQLite数据库查询返回的结果，包含各个指标的值。
    - excel_to_sqlite_indicator_map (dict): Excel指标名称到SQLite字段的映射字典。
    - config (dict): 包含样式配置的字典。

    返回:
    - None: 此函数没有返回值。
    """
    excel_indicators = [cell.value for cell in sheet[config['mapping']['indicator_row']][config['mapping']['column_num']:]]
    indicator_column_map = {indicator: idx + (config['mapping']['column_num'] + 1) for idx, indicator in enumerate(excel_indicators)}

    # excel_indicator excel指标名称   column_idx 指标对应的列号
    for excel_indicator, column_idx in indicator_column_map.items():
        sqlite_field = excel_to_sqlite_indicator_map.get(excel_indicator)
        print(sqlite_field)

        if sqlite_field and query_result:
            result_index = list(excel_to_sqlite_indicator_map.values()).index(sqlite_field)
            cell_value = query_result[result_index]
            cell = sheet.cell(row=project_row, column=column_idx)
            cell.value = cell_value
            set_cell_style(cell, config)

# 主函数
def main():
    # 加载配置文件
    config = load_config()

    # 读取Excel模板
    input_path = config['input_path']
    workbook, sheet, excel_indicators, project_map = read_excel_template(input_path, config)

    # 从配置文件中获取Excel指标名称到SQLite字段的映射
    excel_to_sqlite_indicator_map = config['indicator_mapping']
    sqlite_indicators = list(excel_to_sqlite_indicator_map.values())

    # 连接数据库，只执行一次
    conn = sqlite3.connect(config['sqlite']['db_path'])

    # 遍历Excel中的项目名称
    for project_name, excel_row in project_map.items():
        project_id = config['project_mapping'].get(project_name)
        if project_id:
            # 查询SQLite数据库中的数据
            query_result = query_sqlite_data(conn, project_id, sqlite_indicators, config)
            if query_result:
                # 将查询结果写入Excel
                write_to_excel(sheet, excel_row, query_result, excel_to_sqlite_indicator_map, config)

    # 关闭数据库连接
    conn.close()

    # 保存Excel文件
    output_path = config['output_path']
    workbook.save(output_path)
    print(f"数据已成功写入 {output_path}")

if __name__ == "__main__":
    main()
