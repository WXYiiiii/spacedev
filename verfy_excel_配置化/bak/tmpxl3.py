import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sqlite3

"""
根据维度匹配数据要写入的行
匹配行和列
添加控制excel格式的配置 写入的部分
"""

# 读取配置文件
def load_config(config_path='config.yaml'):
    with open(config_path, 'r', encoding='utf-8') as file:
        return yaml.safe_load(file)

# 连接SQLite，执行查询并获取结果
def query_sqlite_data(project_id, sqlite_indicators, config):
    conn = sqlite3.connect(config['sqlite']['db_path'])
    cursor = conn.cursor()

    # 构建查询语句，选择项目ID和各指标
    query = f"""
    SELECT {', '.join(sqlite_indicators)}
    FROM project_data
    WHERE {config['mapping']['project_id_column']} = ?
    """
    cursor.execute(query, (project_id,))

    # 获取查询结果
    row = cursor.fetchone()

    conn.close()
    return row

# 读取Excel模板，获取项目名称、指标名称的映射关系
def read_excel_template(template_path, config):
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active

    # 获取Excel中的指标字段（第4行），从C列开始
    excel_indicators = [cell.value for cell in sheet[config['mapping']['indicator_row']][2:]]

    # 获取项目名称和对应的Excel行号
    project_map = {}
    for row in range(config['mapping']['start_row'], sheet.max_row + 1):
        project_name = sheet[f"{config['mapping']['project_column']}{row}"].value
        if project_name:
            project_map[project_name] = row

    return workbook, sheet, excel_indicators, project_map

# 设置单元格样式的函数
def set_cell_style(cell, config):
    # 设置字体
    font = Font(name=config['excel_style']['font']['name'],
                size=config['excel_style']['font']['size'],
                bold=config['excel_style']['font']['bold'],
                italic=config['excel_style']['font']['italic'],
                color=config['excel_style']['font']['color'])
    cell.font = font

    # 设置背景颜色
    fill = PatternFill(start_color=config['excel_style']['fill']['start_color'],
                       end_color=config['excel_style']['fill']['end_color'],
                       fill_type=config['excel_style']['fill']['fill_type'])
    cell.fill = fill

    # 设置对齐方式
    alignment = Alignment(horizontal=config['excel_style']['alignment']['horizontal'],
                          vertical=config['excel_style']['alignment']['vertical'])
    cell.alignment = alignment

# 将查询结果写入Excel
def write_to_excel(sheet, project_row, query_result, excel_to_sqlite_indicator_map, config):
    # 获取Excel中指标字段的顺序
    excel_indicators = [cell.value for cell in sheet[config['mapping']['indicator_row']][2:]]

    # 创建一个指标到列号的映射
    indicator_column_map = {indicator: idx + 3 for idx, indicator in enumerate(excel_indicators)}

    # 遍历Excel中的指标名称，并根据配置文件中的映射将数据写入相应的单元格
    for excel_indicator, column_idx in indicator_column_map.items():
        sqlite_field = excel_to_sqlite_indicator_map.get(excel_indicator)
        if sqlite_field:
            # 查询结果列表的索引对应数据库字段
            result_index = list(excel_to_sqlite_indicator_map.values()).index(sqlite_field)
            cell_value = query_result[result_index]
            cell = sheet.cell(row=project_row, column=column_idx)
            cell.value = cell_value

            # 设置单元格样式
            set_cell_style(cell, config)

# 主函数
def main():
    config = load_config()

    # 读取Excel模板
    template_path = 'template.xlsx'
    workbook, sheet, excel_indicators, project_map = read_excel_template(template_path, config)

    # 从配置文件中获取Excel指标名称到SQLite字段的映射
    excel_to_sqlite_indicator_map = config['indicator_mapping']

    # 获取对应的SQLite字段名
    sqlite_indicators = list(excel_to_sqlite_indicator_map.values())

    # 遍历Excel中的项目名称
    for project_name, excel_row in project_map.items():
        # 根据项目名称获取对应的项目ID
        project_id = config['project_mapping'].get(project_name)

        if project_id:
            # 查询SQLite数据库中的数据
            query_result = query_sqlite_data(project_id, sqlite_indicators, config)

            if query_result:
                # 将查询结果写入Excel
                write_to_excel(sheet, excel_row, query_result, excel_to_sqlite_indicator_map, config)

    # 保存Excel文件
    output_path = 'output.xlsx'
    workbook.save(output_path)
    print(f"数据已成功写入 {output_path}")

if __name__ == "__main__":
    main()
