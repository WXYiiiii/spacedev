import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sqlite3
import logging

# 配置日志记录
logging.basicConfig(filename='script.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def log_info(message):
    """记录信息级别日志"""
    logging.info(message)

def log_warning(message):
    """记录警告级别日志"""
    logging.warning(message)

def log_error(message):
    """记录错误级别日志"""
    logging.error(message)

def load_config(config_path='config.yaml'):
    """加载配置文件"""
    with open(config_path, 'r', encoding='utf-8') as file:
        return yaml.safe_load(file)

def query_sqlite_data(project_id, sqlite_indicators, config):
    """
    查询SQLite数据库中的数据
    :param project_id: 项目的ID
    :param sqlite_indicators: 需要查询的指标字段
    :param config: 配置文件
    :return: 查询结果行
    """
    try:
        conn = sqlite3.connect(config['sqlite']['db_path'])
        cursor = conn.cursor()
        query = f"""
        SELECT {', '.join(sqlite_indicators)}
        FROM project_data
        WHERE {config['mapping']['project_id_column']} = ?
        """
        cursor.execute(query, (project_id,))
        row = cursor.fetchone()
        conn.close()
        return row
    except sqlite3.Error as e:
        log_error(f"SQLite查询失败: {str(e)}")
        return None

def read_excel_template(template_path, config):
    """
    读取Excel模板文件并解析项目名称与指标
    :param template_path: Excel模板文件路径
    :param config: 配置文件
    :return: workbook, sheet, excel_indicators, project_map
    """
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active

    # 获取Excel中的指标字段（第4行），从C列开始
    excel_indicators = [cell.value for cell in sheet[config['mapping']['indicator_row']][2:]]

    # 获取项目名称和对应的Excel行号
    project_map = {}
    for row in range(config['mapping']['start_row'], sheet.max_row + 1):
        project_name = sheet[f"{config['mapping']['project_column']}{row}"].value
        if project_name:
            project_map[project_name] = row

    return workbook, sheet, excel_indicators, project_map

def set_cell_style(cell, style_config):
    """
    设置单元格的样式
    :param cell: 要设置样式的单元格
    :param style_config: 样式配置
    """
    # 设置字体
    font = Font(name=style_config['font']['name'],
                size=style_config['font']['size'],
                bold=style_config['font']['bold'],
                italic=style_config['font']['italic'],
                color=style_config['font']['color'])
    cell.font = font

    # 设置背景颜色
    fill = PatternFill(start_color=style_config['fill']['start_color'])
    cell.fill = fill

    # 设置对齐方式
    alignment = Alignment(horizontal=style_config['alignment']['horizontal'],
                          vertical=style_config['alignment']['vertical'])
    cell.alignment = alignment

def apply_formatting(sheet, config):
    """
    根据配置文件应用格式到特定行、列或单元格
    :param sheet: Excel工作表
    :param config: 配置文件
    """
    if 'formatting_targets' in config:
        formatting_targets = config['formatting_targets']
        for target, style_config in formatting_targets.items():
            if target.isdigit():  # 格式化行
                row = int(target)
                for cell in sheet[row]:
                    set_cell_style(cell, style_config)
            elif len(target) == 1:  # 格式化列
                col = target.upper()
                for row in range(1, sheet.max_row + 1):
                    cell = sheet[f"{col}{row}"]
                    set_cell_style(cell, style_config)
            else:  # 格式化单元格
                cell = sheet[target.upper()]
                set_cell_style(cell, style_config)
    else:
        log_info("没有格式化配置，跳过格式化步骤")

def write_to_excel(sheet, project_row, query_result, excel_to_sqlite_indicator_map, config):
    """
    将查询结果写入Excel文件
    :param sheet: Excel工作表
    :param project_row: 项目对应的行号
    :param query_result: 从数据库中查询的结果
    :param excel_to_sqlite_indicator_map: Excel指标与SQLite字段的映射
    :param config: 配置文件
    """
    # 获取Excel中指标字段的顺序
    excel_indicators = [cell.value for cell in sheet[config['mapping']['indicator_row']][2:]]

    # 创建一个指标到列号的映射
    indicator_column_map = {indicator: idx + 3 for idx, indicator in enumerate(excel_indicators)}

    # 遍历Excel中的指标名称，并根据配置文件中的映射将数据写入相应的单元格
    for excel_indicator, column_idx in indicator_column_map.items():
        sqlite_field = excel_to_sqlite_indicator_map.get(excel_indicator)
        if sqlite_field:
            # 查询结果列表的索引对应数据库字段
            result_index = list(excel_to_sqlite_indicator_map.values()).index(sqlite_field)
            cell_value = query_result[result_index]
            cell = sheet.cell(row=project_row, column=column_idx)
            cell.value = cell_value

            # 设置单元格样式
            set_cell_style(cell, config['excel_style'])  # 应用默认样式

def process_excel_files(input_files, output_files, config):
    """
    处理多个输入和输出Excel文件
    :param input_files: 输入Excel文件路径列表
    :param output_files: 输出Excel文件路径列表
    :param config: 配置文件
    """
    for input_file, output_file in zip(input_files, output_files):
        try:
            log_info(f"处理文件 {input_file}")
            workbook, sheet, excel_indicators, project_map = read_excel_template(input_file, config)
            excel_to_sqlite_indicator_map = config['indicator_mapping']
            sqlite_indicators = list(excel_to_sqlite_indicator_map.values())
            for project_name, excel_row in project_map.items():
                project_id = config['project_mapping'].get(project_name)
                if project_id:
                    query_result = query_sqlite_data(project_id, sqlite_indicators, config)
                    if query_result:
                        write_to_excel(sheet, excel_row, query_result, excel_to_sqlite_indicator_map, config)
                        log_info(f"项目 {project_name} 的数据已写入 Excel")
                    else:
                        log_warning(f"项目 {project_name} 的查询结果为空")
                else:
                    log_warning(f"项目 {project_name} 的项目ID未找到")
            apply_formatting(sheet, config)
            workbook.save(output_file)
            log_info(f"数据已成功写入 {output_file}")
        except Exception as e:
            log_error(f"处理文件 {input_file} 失败: {str(e)}")

def main():
    """主函数，加载配置并处理Excel文件"""
    config = load_config()
    input_files = config['files']['input_files']
    output_files = config['files']['output_files']
    process_excel_files(input_files, output_files, config)

if __name__ == "__main__":
    main()
