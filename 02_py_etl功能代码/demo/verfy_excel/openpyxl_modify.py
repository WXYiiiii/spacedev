from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 加載現有的Excel文件
wb = load_workbook('./data/input.xlsx')
sheet = wb.active

# 讀取特定單元格的值並進行修改
cell = sheet['A1']
cell.value = 'Modified Value'

# 設置字體
cell.font = Font(size=14, bold=True, color='FF0000')

# 設置背景顏色
cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# 設置列寬
sheet.column_dimensions['A'].width = 20

# 居中對齊
cell.alignment = Alignment(horizontal='center', vertical='center')

# 將修改後的Excel保存回原文件
wb.save('./data/pyxl_modified.xlsx')