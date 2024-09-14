from csv import excel

import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sqlite3

"""
根据维度匹配数据要写入的行
匹配行和列
添加控制excel格式的配置 写入的部分
"""

# 读取配置文件
def load_config(config_path='config.yaml'):
    """
    从指定路径加载YAML配置文件并返回其内容。

    Args:
        config_path (str): 配置文件的路径，默认为 'config.yaml'。

    Returns:
        dict: 配置文件的内容，以字典形式返回。
    """
    with open(config_path, 'r', encoding='utf-8') as file:
        return yaml.safe_load(file)    # 将yaml中的配置文件返回为字典格式




# 连接SQLite，执行查询并获取结果
def query_sqlite_data(project_id, sqlite_indicators, config):
    """
    根据项目ID（就是维度id）从SQLite数据库中查询对应的指标数据。

    Args:
        project_id (str): 项目的唯一标识符。
        sqlite_indicators (list): 要查询的指标列表。  表中的字段名
        config (dict): 配置文件内容。

    Returns:
        tuple: 查询结果的第一行数据，若无结果则返回 None。
    """
    # 连接到SQLite数据库
    conn = sqlite3.connect(config['sqlite']['db_path'])        # 这里取的是字典（配置文件）里的值
    cursor = conn.cursor()

    # 构建查询语句，选择项目ID和各指标
    query = f"""
    SELECT {', '.join(sqlite_indicators)}
    FROM project_data
    WHERE {config['mapping']['project_id_column']} = ?        
    """
    # 这里的 config['mapping']['project_id_column']} 是配置中的列名称。
    # ？是一个占位符，防止SQL注入   通过这个where语句 匹配excel中对应的行

    # 执行查询
    cursor.execute(query, (project_id,))
    # （project_id,） 是一个只有一个元素的元组   单个元素的元组需要在元素后面加上逗号
    # project_id 是替换？用的，它仍然是一个变量



    # 获取查询结果
    row = cursor.fetchone()

    # print(row)
    # row是元组格式，这里是数据表中的一行数据

    # 关闭数据库连接
    conn.close()
    return row



# 设置单元格样式的函数
def set_cell_style(cell, config):
    """
    设置给定单元格的样式，包括字体、背景颜色和对齐方式。

    Args:
        cell (Cell): 需要设置样式的单元格。
        config (dict): 配置文件内容。
    """
    # 设置字体
    font = Font(name=config['excel_style']['font']['name'],
                size=config['excel_style']['font']['size'],
                bold=config['excel_style']['font']['bold'],
                italic=config['excel_style']['font']['italic'],
                color=config['excel_style']['font']['color'])
    cell.font = font

    # 设置背景颜色
    fill = PatternFill(start_color=config['excel_style']['fill']['start_color']

                       )
    cell.fill = fill

    # 设置对齐方式
    alignment = Alignment(horizontal=config['excel_style']['alignment']['horizontal'],
                          vertical=config['excel_style']['alignment']['vertical'])
    cell.alignment = alignment




# 读取Excel模板，获取项目名称、指标名称的映射关系
def read_excel_template(template_path, config):
    """
    # workbook对象 -> sheet --> 指标字段 | 维度和excel行号的字典
    读取Excel模板文件，并提取项目名称和指标名称的映射关系。

    Args:
        template_path (str): Excel模板文件的路径。
        config (dict): 配置文件内容。

    Returns:
        tuple: 包含工作簿、工作表、指标列表和项目名称映射的元组。
    """
    # 加载Excel工作簿
    workbook = openpyxl.load_workbook(template_path)

    sheet = workbook.active
    # sheet = workbook['Sheet1']  # 假设工作表名称为 'Sheet1'


    # 获取Excel中的指标字段（第4行），从C列开始
    excel_indicators = [cell.value for cell in sheet[config['mapping']['indicator_row']] [config['mapping']['column_num']:]  ]

    # print(excel_indicators)

    # 获取维度名称和对应的Excel行号
    project_map = {}    # 将维度名称和excel的行号放到字典中
    for row in range(config['mapping']['start_row'], sheet.max_row + 1):                          # 循环的行数 指标所在的行到sheet页面的最后一行
        project_name = sheet[f"{config['mapping']['project_column']}{row}"].value
        # print(project_name)


        if project_name:
            project_map[project_name] = row


    # print(workbook,sheet,excel_indicators,project_map)

    return workbook, sheet, excel_indicators, project_map




# 将查询结果写入Excel
def write_to_excel(sheet, project_row, query_result, excel_to_sqlite_indicator_map, config):
    """
    将从SQLite查询得到的结果写入指定的Excel工作表。

    Args:
        sheet (Worksheet): 要写入的Excel工作表。
        project_row (int): 维度在Excel中的行号。
        query_result (tuple): 从SQLite查询得到的结果。
        excel_to_sqlite_indicator_map (dict): Excel指标名称到SQLite字段的映射。
        config (dict): 配置文件内容。
    """
    # 获取Excel中指标字段的顺序
    # excel_indicators = [cell.value for cell in sheet[config['mapping']['indicator_row']][2:]]

    excel_indicators = [cell.value for cell in sheet[config['mapping']['indicator_row']] [config['mapping']['column_num']:]  ]
    # print(excel_indicators)

    # 创建一个指标到列号的映射
    indicator_column_map = {indicator: idx + 3 for idx, indicator in enumerate(excel_indicators)}
    # print(indicator_column_map)


    print(excel_to_sqlite_indicator_map)

    # 遍历Excel中的指标名称，并根据配置文件中的映射将数据写入相应的单元格
    for excel_indicator, column_idx in indicator_column_map.items():               # 遍历指标和excel的列号的键值对
        sqlite_field = excel_to_sqlite_indicator_map.get(excel_indicator)           #
        if sqlite_field:
            # 查询结果列表的索引对应数据库字段
            result_index = list(excel_to_sqlite_indicator_map.values()).index(sqlite_field)
            cell_value = query_result[result_index]
            cell = sheet.cell(row=project_row, column=column_idx)
            cell.value = cell_value

            # 设置单元格样式
            set_cell_style(cell, config)

# 主函数
def main():
    """
    主程序入口，负责加载配置、读取Excel模板、查询数据库并写入结果。
    """
    # 加载配置文件
    config = load_config()

    # 读取Excel模板
    template_path = './data/input_file1.xlsx'
    workbook, sheet, excel_indicators, project_map = read_excel_template(template_path, config)

    # 从配置文件中获取Excel指标名称到SQLite字段的映射
    excel_to_sqlite_indicator_map = config['indicator_mapping']

    # 获取对应的SQLite字段名
    sqlite_indicators = list(excel_to_sqlite_indicator_map.values())

    # 遍历Excel中的项目名称
    for project_name, excel_row in project_map.items():
        # 根据项目名称获取对应的项目ID
        project_id = config['project_mapping'].get(project_name)

        if project_id:
            # 查询SQLite数据库中的数据
            # print(excel_row)
            # print(project_id)
            query_result = query_sqlite_data(project_id, sqlite_indicators, config)

            if query_result:
                # 将查询结果写入Excel
                write_to_excel(sheet, excel_row, query_result, excel_to_sqlite_indicator_map, config)

    # 保存Excel文件
    output_path = 'output.xlsx'
    workbook.save(output_path)
    print(f"数据已成功写入 {output_path}")

if __name__ == "__main__":
    main()