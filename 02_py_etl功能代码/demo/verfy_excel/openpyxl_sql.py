import pandas as pd
from openpyxl import load_workbook

# 模擬連接到數據庫並獲取數據
new_data = {
    'Name': ['David', 'Eve'],
    'Age': [40, 45],
    'City': ['Miami', 'Seattle']
}

# 轉換新數據為DataFrame
new_df = pd.DataFrame(new_data)

# 加載Excel文件
wb = load_workbook('./data/input.xlsx')
sheet = wb.active

# 獲取現有數據
existing_data = []
for row in sheet.iter_rows(values_only=True):
    existing_data.append(row)

# 獲取現有數據的列名
header = existing_data[0]

# 構建新的數據列表，將新數據插入現有數據之後
new_data_rows = [header] + existing_data[1:] + new_df.values.tolist()

# 清空工作表
sheet.delete_rows(1, sheet.max_row)

# 將新的數據寫入Excel文件
for row_data in new_data_rows:
    sheet.append(row_data)

# 保存更新後的Excel文件
wb.save('./data/pyxl_updated.xlsx')