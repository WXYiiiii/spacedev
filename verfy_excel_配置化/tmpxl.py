import yaml
import openpyxl
import sqlite3
import pandas as pd

"""
匹配列名
"""

# 读取配置文件
def load_config(config_path='config.yaml'):
    with open(config_path, 'r', encoding='utf-8') as file:
        return yaml.safe_load(file)

# 连接SQLite，执行查询并获取结果
def query_sqlite_data(config, sqlite_indicators):
    conn = sqlite3.connect(config['sqlite']['db_path'])
    cursor = conn.cursor()

    # 构建查询语句，选择项目ID和各指标
    query = f"""
    SELECT {config['mapping']['project_id_column']}, {', '.join(sqlite_indicators)}
    FROM project_data
    """
    cursor.execute(query)

    # 获取查询结果并转换为DataFrame
    columns = [config['mapping']['project_id_column']] + sqlite_indicators
    rows = cursor.fetchall()
    df = pd.DataFrame(rows, columns=columns)

    conn.close()
    return df

# 读取Excel模板，获取项目名称、指标名称的映射关系
def read_excel_template(template_path, config):
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active

    # 获取Excel中的指标字段（第4行），从C列开始
    excel_indicators = [cell.value for cell in sheet[config['mapping']['indicator_row']][2:]]

    # 获取项目名称和对应的Excel行号
    project_map = {}
    for row in range(config['mapping']['start_row'], sheet.max_row + 1):
        project_name = sheet[f"{config['mapping']['project_column']}{row}"].value
        if project_name:
            project_map[project_name] = row

    return workbook, sheet, excel_indicators, project_map

# 将查询结果写入Excel
def write_to_excel(sheet, df, project_map, excel_to_sqlite_indicator_map, config):
    for _, row in df.iterrows():
        project_id = row[config['mapping']['project_id_column']]

        # 根据项目ID获取项目名称
        project_name = next((name for name, pid in config['project_mapping'].items() if pid == project_id), None)
        if project_name and project_name in project_map:
            excel_row = project_map[project_name]

            # 从C列开始写入指标数据
            for col_idx, excel_indicator in enumerate(excel_to_sqlite_indicator_map.keys(), start=3):
                sqlite_field = excel_to_sqlite_indicator_map[excel_indicator]
                cell = sheet.cell(row=excel_row, column=col_idx)
                cell.value = row[sqlite_field]

# 主函数
def main():
    config = load_config()

    # 读取Excel模板
    template_path = 'template.xlsx'
    workbook, sheet, excel_indicators, project_map = read_excel_template(template_path, config)

    # 从配置文件中获取Excel指标名称到SQLite字段的映射
    excel_to_sqlite_indicator_map = config['indicator_mapping']

    # 获取对应的SQLite字段名
    sqlite_indicators = list(excel_to_sqlite_indicator_map.values())

    # 查询SQLite数据
    df = query_sqlite_data(config, sqlite_indicators)

    # 将数据写入Excel
    write_to_excel(sheet, df, project_map, excel_to_sqlite_indicator_map, config)

    # 保存Excel文件
    output_path = 'output.xlsx'
    workbook.save(output_path)
    print(f"数据已成功写入 {output_path}")

if __name__ == "__main__":
    main()
