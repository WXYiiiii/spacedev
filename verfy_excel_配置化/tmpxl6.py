import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sqlite3
import logging

logging.basicConfig(filename='script.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def log_info(message):
    logging.info(message)

def log_warning(message):
    logging.warning(message)

def log_error(message):
    logging.error(message)

def load_config(config_path='config.yaml'):
    with open(config_path, 'r', encoding='utf-8') as file:
        return yaml.safe_load(file)

def query_sqlite_data(project_id, sqlite_indicators, config):
    try:
        conn = sqlite3.connect(config['sqlite']['db_path'])
        cursor = conn.cursor()
        query = f"""
        SELECT {', '.join(sqlite_indicators)}
        FROM project_data
        WHERE {config['mapping']['project_id_column']} = ?
        """
        cursor.execute(query, (project_id,))
        row = cursor.fetchone()
        conn.close()
        return row
    except sqlite3.Error as e:
        log_error(f"SQLite查询失败: {str(e)}")
        return None

def read_excel_template(template_path, config):
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active

    excel_indicators = [cell.value for cell in sheet[config['mapping']['indicator_row']][2:]]

    project_map = {}
    for row in range(config['mapping']['start_row'], sheet.max_row + 1):
        project_name = sheet[f"{config['mapping']['project_column']}{row}"].value
        if project_name:
            project_map[project_name] = row

    return workbook, sheet, excel_indicators, project_map

def set_cell_style(cell, style_config):
    font = Font(name=style_config['font']['name'],
                size=style_config['font']['size'],
                bold=style_config['font']['bold'],
                italic=style_config['font']['italic'],
                color=style_config['font']['color'])
    cell.font = font

    fill = PatternFill(start_color=style_config['fill']['start_color'])
    cell.fill = fill

    alignment = Alignment(horizontal=style_config['alignment']['horizontal'],
                          vertical=style_config['alignment']['vertical'])
    cell.alignment = alignment

def apply_formatting(sheet, config):
    if 'formatting_targets' in config:
        formatting_targets = config['formatting_targets']
        for target, style_config in formatting_targets.items():
            if target.isdigit():
                row = int(target)
                for cell in sheet[row]:
                    set_cell_style(cell, style_config)
            elif len(target) == 1:
                col = target.upper()
                for row in range(1, sheet.max_row + 1):
                    cell = sheet[f"{col}{row}"]
                    set_cell_style(cell, style_config)
            else:
                cell = sheet[target.upper()]
                set_cell_style(cell, style_config)
    else:
        log_info("没有格式化配置，跳过格式化步骤")

def write_to_excel(sheet, project_row, query_result, excel_to_sqlite_indicator_map, config):
    excel_indicators = [cell.value for cell in sheet[config['mapping']['indicator_row']][2:]]

    indicator_column_map = {indicator: idx + 3 for idx, indicator in enumerate(excel_indicators)}

    for excel_indicator, column_idx in indicator_column_map.items():
        sqlite_field = excel_to_sqlite_indicator_map.get(excel_indicator)
        if sqlite_field:
            result_index = list(excel_to_sqlite_indicator_map.values()).index(sqlite_field)
            cell_value = query_result[result_index]
            cell = sheet.cell(row=project_row, column=column_idx)
            cell.value = cell_value

            set_cell_style(cell, config['excel_style'])

def process_excel_files(input_files, output_files, config):
    for input_file, output_file in zip(input_files, output_files):
        try:
            log_info(f"处理文件 {input_file}")
            workbook, sheet, excel_indicators, project_map = read_excel_template(input_file, config)
            excel_to_sqlite_indicator_map = config['indicator_mapping']
            sqlite_indicators = list(excel_to_sqlite_indicator_map.values())
            for project_name, excel_row in project_map.items():
                project_id = config['project_mapping'].get(project_name)
                if project_id:
                    query_result = query_sqlite_data(project_id, sqlite_indicators, config)
                    if query_result:
                        write_to_excel(sheet, excel_row, query_result, excel_to_sqlite_indicator_map, config)
                        log_info(f"项目 {project_name} 的数据已写入 Excel")
                    else:
                        log_warning(f"项目 {project_name} 的查询结果为空")
                else:
                    log_warning(f"项目 {project_name} 的项目ID未找到")
            apply_formatting(sheet, config)
            workbook.save(output_file)
            log_info(f"数据已成功写入 {output_file}")
        except Exception as e:
            log_error(f"处理文件 {input_file} 失败: {str(e)}")

def main():
    config = load_config()
    input_files = config['files']['input_files']
    output_files = config['files']['output_files']
    process_excel_files(input_files, output_files, config)

if __name__ == "__main__":
    main()
