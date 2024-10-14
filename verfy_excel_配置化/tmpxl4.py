import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sqlite3

"""
根据维度匹配数据要写入的行
匹配行和列
添加控制excel格式的配置 写入的部分
添加控制特定单元格格式

- 没有配置输入输出路径
"""


# 读取配置文件
def load_config(config_path='config.yaml'):
    with open(config_path, 'r', encoding='utf-8') as file:
        config = yaml.safe_load(file)
    return config

# 连接SQLite，执行查询并获取结果
def query_sqlite_data(project_id, sqlite_indicators, config):
    conn = sqlite3.connect(config['sqlite']['db_path'])
    cursor = conn.cursor()

    # 构建查询语句，选择项目ID和各指标
    query = f"""
    SELECT {', '.join(sqlite_indicators)}
    FROM project_data
    WHERE {config['mapping']['project_id_column']} = ?
    """
    cursor.execute(query, (project_id,))

    # 获取查询结果
    row = cursor.fetchone()

    conn.close()
    return row

# 读取Excel模板，获取项目名称、指标名称的映射关系
def read_excel_template(template_path, config):
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active

    # 获取Excel中的指标字段（第4行），从C列开始
    excel_indicators = [cell.value for cell in sheet[config['mapping']['indicator_row']][2:]]

    # 获取项目名称和对应的Excel行号
    project_map = {}
    for row in range(config['mapping']['start_row'], sheet.max_row + 1):
        project_name = sheet[f"{config['mapping']['project_column']}{row}"].value
        if project_name:
            project_map[project_name] = row

    return workbook, sheet, excel_indicators, project_map

# 设置单元格样式的函数
def set_cell_style(cell, style_config):
    # 设置字体
    font = Font(name=style_config['font']['name'],
                size=style_config['font']['size'],
                bold=style_config['font']['bold'],
                italic=style_config['font']['italic'],
                color=style_config['font']['color'])
    cell.font = font

    # 设置背景颜色
    fill = PatternFill(start_color=style_config['fill']['start_color'])
    cell.fill = fill

    # 设置对齐方式
    alignment = Alignment(horizontal=style_config['alignment']['horizontal'],
                          vertical=style_config['alignment']['vertical'])
    cell.alignment = alignment

# 应用格式到特定行、列或单元格
# def apply_formatting(sheet, config):
#
#
#     formatting_targets = config.get('formatting_targets',{})
#     for target, style_config in formatting_targets.items():
#         if isinstance(target, str):
#             if target.isdigit():  # 处理行格式
#                 row = int(target)
#                 for cell in sheet[row]:
#                     set_cell_style(cell, style_config)
#             elif len(target) == 1:  # 处理列格式
#                 col = target.upper()
#                 for row in range(1, sheet.max_row + 1):
#                     cell = sheet[f"{col}{row}"]
#                     set_cell_style(cell, style_config)
#             else:  # 处理具体单元格
#                 cell = sheet[target.upper()]
#                 set_cell_style(cell, style_config)

# 将查询结果写入Excel
def write_to_excel(sheet, project_row, query_result, excel_to_sqlite_indicator_map, config):
    # 获取Excel中指标字段的顺序
    excel_indicators = [cell.value for cell in sheet[config['mapping']['indicator_row']][2:]]

    # 创建一个指标到列号的映射
    indicator_column_map = {indicator: idx + 3 for idx, indicator in enumerate(excel_indicators)}

    print(indicator_column_map)


    # 遍历Excel中的指标名称，并根据配置文件中的映射将数据写入相应的单元格
    for excel_indicator, column_idx in indicator_column_map.items():
        sqlite_field = excel_to_sqlite_indicator_map.get(excel_indicator)
        if sqlite_field:
            # 查询结果列表的索引对应数据库字段
            result_index = list(excel_to_sqlite_indicator_map.values()).index(sqlite_field)
            cell_value = query_result[result_index]
            cell = sheet.cell(row=project_row, column=column_idx)
            cell.value = cell_value

            # 设置单元格样式
            set_cell_style(cell, config['excel_style'])  # 应用默认样式

# 主函数
def main():
    config = load_config()

    # 读取Excel模板
    template_path = './data/input_file1.xlsx'
    workbook, sheet, excel_indicators, project_map = read_excel_template(template_path, config)

    # 从配置文件中获取Excel指标名称到SQLite字段的映射
    excel_to_sqlite_indicator_map = config['indicator_mapping']

    # 获取对应的SQLite字段名
    sqlite_indicators = list(excel_to_sqlite_indicator_map.values())

    # 遍历Excel中的项目名称
    for project_name, excel_row in project_map.items():
        # 根据项目名称获取对应的项目ID
        project_id = config['project_mapping'].get(project_name)

        if project_id:
            # 查询SQLite数据库中的数据
            query_result = query_sqlite_data(project_id, sqlite_indicators, config)

            if query_result:
                # 将查询结果写入Excel
                write_to_excel(sheet, excel_row, query_result, excel_to_sqlite_indicator_map, config)

    # 应用用户指定的格式到特定行、列或单元格
    apply_formatting(sheet, config)

    # 保存Excel文件
    output_path = 'output.xlsx'
    workbook.save(output_path)
    print(f"数据已成功写入 {output_path}")

if __name__ == "__main__":
    main()
