import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sqlite3


# 读取配置文件
def load_config(config_path='config.yaml'):
    """
    从指定路径加载YAML配置文件并返回其内容。

    Args:
        config_path (str): 配置文件的路径，默认为 'config.yaml'。

    Returns:
        dict: 解析后的配置文件内容，以字典形式返回。
    """
    with open(config_path, 'r', encoding='utf-8') as file:
        return yaml.safe_load(file)


# 连接SQLite，执行查询并获取结果
def query_sqlite_data(conn, project_id, sqlite_indicators, config):
    """
    根据项目ID从SQLite数据库中查询对应的指标数据。

    Args:
        conn (sqlite3.Connection): 已连接的SQLite数据库连接对象。
        project_id (str): 项目的唯一标识符，用来匹配数据库中的项目记录。
        sqlite_indicators (list): 要查询的数据库字段名称列表。
        config (dict): 从配置文件加载的配置信息。

    Returns:
        tuple: 查询结果的第一行数据，若无结果则返回 None。
    """
    cursor = conn.cursor()

    # 构建SQL查询语句，查询项目的指定指标
    query = f"""
    SELECT {', '.join(sqlite_indicators)}  
    FROM project_data                      
    WHERE {config['mapping']['project_id_column']} = ?  
    """

    # 执行查询，传入项目ID，获取查询结果
    cursor.execute(query, (project_id,))
    return cursor.fetchone()  # 获取第一行数据，若无数据则返回None


# 设置单元格样式的函数
def set_cell_style(cell, config):
    """
    根据配置文件中的格式设置单元格样式，包括字体、填充色和对齐方式。

    Args:
        cell (Cell): 要设置样式的Excel单元格对象。
        config (dict): 配置文件内容。
    """
    # 设置字体样式，包括名称、大小、粗体、斜体、颜色等
    cell.font = Font(
        name=config['excel_style']['font']['name'],
        size=config['excel_style']['font']['size'],
        bold=config['excel_style']['font']['bold'],
        italic=config['excel_style']['font']['italic'],
        color=config['excel_style']['font']['color']
    )

    # 设置单元格背景颜色
    cell.fill = PatternFill(start_color=config['excel_style']['fill']['start_color'])

    # 设置单元格的对齐方式
    cell.alignment = Alignment(
        horizontal=config['excel_style']['alignment']['horizontal'],
        vertical=config['excel_style']['alignment']['vertical']
    )


# 读取Excel模板，获取项目名称、指标名称的映射关系
def read_excel_template(template_path, config):
    """
    读取Excel模板文件，提取项目名称和指标名称的映射关系。

    Args:
        template_path (str): Excel模板文件的路径。
        config (dict): 配置文件内容。

    Returns:
        tuple: 返回包含工作簿对象、工作表对象、指标列表、项目名称与行号映射的元组。
    """
    # 加载Excel工作簿
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active  # 获取活动的工作表

    # 获取Excel中的指标字段列表，从指定列开始
    excel_indicators = [cell.value for cell in
                        sheet[config['mapping']['indicator_row']][config['mapping']['column_num']:]]

    # 构建维度名称与对应的行号的映射
    project_map = {
        sheet[f"{config['mapping']['project_column']}{row}"].value: row  # 将项目名称和对应行号作为键值对存储
        for row in range(config['mapping']['start_row'], sheet.max_row + 1)  # 从指定起始行遍历到最后一行
        if sheet[f"{config['mapping']['project_column']}{row}"].value  # 确保项目名称不为空
    }

    return workbook, sheet, excel_indicators, project_map


# 将查询结果写入Excel
def write_to_excel(sheet, project_row, query_result, excel_to_sqlite_indicator_map, config):
    """
    将从SQLite查询得到的结果写入指定的Excel工作表。

    Args:
        sheet (Worksheet): 要写入的Excel工作表对象。
        project_row (int): 维度在Excel中的行号，对应于查询结果的行。
        query_result (tuple): 从SQLite查询得到的结果。
        excel_to_sqlite_indicator_map (dict): Excel指标名称到SQLite字段的映射。
        config (dict): 配置文件内容。
    """
    # 获取Excel中所有的指标名称和列号
    excel_indicators = [cell.value for cell in
                        sheet[config['mapping']['indicator_row']][config['mapping']['column_num']:]]

    # 构建指标名称到Excel列号的映射
    indicator_column_map = {indicator: idx + config['mapping']['column_num'] for idx, indicator in
                            enumerate(excel_indicators)}

    # 遍历每个Excel指标，匹配SQLite数据字段并写入相应单元格
    for excel_indicator, column_idx in indicator_column_map.items():
        sqlite_field = excel_to_sqlite_indicator_map.get(excel_indicator)  # 获取SQLite中的字段名称
        if sqlite_field and query_result:  # 确保查询结果不为空
            result_index = list(excel_to_sqlite_indicator_map.values()).index(sqlite_field)  # 获取该字段的查询结果索引
            cell_value = query_result[result_index]  # 从查询结果中获取相应的值
            cell = sheet.cell(row=project_row, column=column_idx)  # 定位到相应单元格
            cell.value = cell_value  # 将值写入单元格
            set_cell_style(cell, config)  # 应用单元格样式


# 主函数
def main():
    """
    主程序入口，负责加载配置、读取Excel模板、查询数据库并写入结果。
    """
    # 加载配置文件
    config = load_config()

    # 读取Excel模板并获取相关信息
    template_path = './data/input_file1.xlsx'
    workbook, sheet, excel_indicators, project_map = read_excel_template(template_path, config)

    # 获取Excel指标名称到SQLite字段的映射
    excel_to_sqlite_indicator_map = config['indicator_mapping']
    sqlite_indicators = list(excel_to_sqlite_indicator_map.values())  # 从映射中获取所有要查询的数据库字段

    # 连接SQLite数据库（仅执行一次连接）
    conn = sqlite3.connect(config['sqlite']['db_path'])

    # 遍历Excel中的项目名称和对应的行号
    for project_name, excel_row in project_map.items():
        project_id = config['project_mapping'].get(project_name)  # 获取项目名称对应的ID
        if project_id:
            # 查询SQLite数据库中的数据
            query_result = query_sqlite_data(conn, project_id, sqlite_indicators, config)
            if query_result:
                # 将查询结果写入Excel相应的单元格
                write_to_excel(sheet, excel_row, query_result, excel_to_sqlite_indicator_map, config)

    # 关闭数据库连接
    conn.close()

    # 保存修改后的Excel文件
    output_path = 'output.xlsx'
    workbook.save(output_path)
    print(f"数据已成功写入 {output_path}")


# 程序入口
if __name__ == "__main__":
    main()

#test